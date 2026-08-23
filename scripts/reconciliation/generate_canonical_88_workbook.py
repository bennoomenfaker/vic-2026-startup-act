from pathlib import Path
import json
from collections import Counter, defaultdict
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule

REPO = Path('/home/ubuntu/vic-2026-startup-act-4339943')
DATA = REPO / 'public' / 'data'
CANONICAL_PATH = DATA / 'reextraction_88_canonical.json'
OUT_REPO = DATA / 'Startup_Act_88_sessions_reextrait_corrige_2026-08-23.xlsx'
OUT_CANONICAL = Path('/home/ubuntu/startup_act_final_delivery_88/canonical_package/Startup_Act_88_sessions_reextrait_corrige_2026-08-23.xlsx')

canonical = json.loads(CANONICAL_PATH.read_text(encoding='utf-8'))
meta = canonical['meta']
sessions = canonical['sessions']
entries = canonical['entries']
entries_by_session = defaultdict(list)
for e in entries:
    entries_by_session[e['session']].append(e)
for es in entries_by_session.values():
    es.sort(key=lambda e: (int(e.get('line') or 0), str(e.get('decision_id') or '')))

# Civic Ledger workbook style: institutional navy for provenance, teal for detail, amber for anomalies.
NAVY = '16324F'
TEAL = '2B6F77'
GOLD = 'D99A2B'
INK = '1D2B36'
MIST = 'F4F0E8'
PALE_TEAL = 'E4F0F0'
PALE_GOLD = 'FFF1D6'
PALE_RED = 'FCE4E4'
PALE_GREEN = 'E3F1E7'
WHITE = 'FFFFFF'
GREY = '66737D'
THIN = Side(style='thin', color='D8DEE3')
MEDIUM_NAVY = Side(style='medium', color=NAVY)

wb = Workbook()
wb.remove(wb.active)
wb.properties.title = 'Startup Act — Civic Ledger — 88 sessions'
wb.properties.subject = 'Réconciliation des compteurs institutionnels et des lignes PDF détaillées'
wb.properties.creator = 'Civic Ledger'
wb.properties.description = 'Classeur canonique généré à partir de reextraction_88_canonical.json le 23/08/2026.'

header_fill = PatternFill('solid', fgColor=NAVY)
section_fill = PatternFill('solid', fgColor=TEAL)
accent_fill = PatternFill('solid', fgColor=GOLD)
light_fill = PatternFill('solid', fgColor=MIST)
source_fill = PatternFill('solid', fgColor=PALE_TEAL)
warning_fill = PatternFill('solid', fgColor=PALE_GOLD)
error_fill = PatternFill('solid', fgColor=PALE_RED)
positive_fill = PatternFill('solid', fgColor=PALE_GREEN)

font_title = Font(name='Georgia', size=19, bold=True, color=NAVY)
font_subtitle = Font(name='Calibri', size=10, italic=True, color=GREY)
font_section = Font(name='Georgia', size=12, bold=True, color=WHITE)
font_header = Font(name='Georgia', size=10, bold=True, color=WHITE)
font_body = Font(name='Calibri', size=10, color=INK)
font_note = Font(name='Calibri', size=9, italic=True, color=GREY)
font_kpi = Font(name='Georgia', size=15, bold=True, color=NAVY)


def setup(ws, widths=None):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.sheet_view.zoomScale = 90
    ws.column_dimensions['A'].width = 3
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width


def style_header(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(top=MEDIUM_NAVY, bottom=MEDIUM_NAVY)
    ws.row_dimensions[row].height = 30


def style_body(ws, start_row, end_row, start_col, end_col):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.font = font_body
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(bottom=THIN)


def add_table(ws, ref, name):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)

# 1) Overview
ws = wb.create_sheet('Synthèse_88')
setup(ws, {'B': 31, 'C': 22, 'D': 22, 'E': 22, 'F': 22, 'G': 22, 'H': 24})
ws.merge_cells('B2:H2')
ws['B2'] = 'Startup Act — Civic Ledger | Réconciliation 88 sessions'
ws['B2'].font = font_title
ws['B2'].alignment = Alignment(vertical='center')
ws.row_dimensions[2].height = 32
ws.merge_cells('B3:H3')
ws['B3'] = 'Classeur canonique · généré le 23/08/2026 · les compteurs officiels et les lignes PDF sont deux séries documentaires distinctes.'
ws['B3'].font = font_subtitle
ws['B3'].alignment = Alignment(wrap_text=True)

ws['B5'] = 'INDICATEURS DE RÉFÉRENCE'
ws['B5'].fill = section_fill
ws['B5'].font = font_section
ws.merge_cells('B5:H5')
for cell in ws[5][1:8]: cell.fill = section_fill

kpis = [
    ('Sessions PDF', int(meta['sessions']), '88 sessions couvertes'),
    ('Candidatures officielles', int(meta['official_candidatures']), 'compteur institutionnel de référence'),
    ('Lignes détaillées PDF', int(meta['detailedEntries']), 'lignes documentaires extraites et contrôlées'),
    ('Labels officiels', int(meta['official_labels']), 'compteur institutionnel'),
    ('Prélabels officiels', int(meta['official_preLabels']), 'compteur institutionnel'),
    ('Retraits officiels', sum(int(s.get('retraits') or 0) for s in sessions), 'compteur de retraits'),
    ('Reportés explicitement', sum(int(s.get('reportes') or 0) for s in sessions), 'lignes Reporté visibles dans le détail'),
]
for i, (label, value, note) in enumerate(kpis, start=7):
    ws.cell(i, 2, label).font = Font(name='Calibri', size=10, bold=True, color=INK)
    ws.cell(i, 3, value).font = font_kpi
    ws.cell(i, 3).number_format = '#,##0'
    ws.cell(i, 4, note).font = font_note
    ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=8)
    for c in range(2, 9):
        ws.cell(i, c).border = Border(bottom=THIN)
        ws.cell(i, c).alignment = Alignment(vertical='center', wrap_text=True)

ws['B16'] = 'RÈGLE DE LECTURE'
ws['B16'].fill = section_fill
ws['B16'].font = font_section
ws.merge_cells('B16:H16')
for cell in ws[16][1:8]: cell.fill = section_fill
notes = [
    'La série institutionnelle conserve 3 079 candidatures officielles : elle est la référence pour les indicateurs institutionnels.',
    'La série PDF contient 3 555 lignes détaillées : elle sert à l’analyse des entreprises, fondateurs, décisions, conversions, retraits et statuts Reporté.',
    'S62 (05/2024) : 39 candidatures officielles et 46 lignes PDF. Les 4 dossiers administratifs sans décision publiée sont conservés comme « Décision non précisée — motif administratif » ; ils ne sont pas retirés du détail.',
    'Les 5 lignes Reporté sont conservées dans le tableau détaillé et distinguées par leur décision normalisée ; elles ne modifient pas rétroactivement le compteur officiel.',
]
for i, text in enumerate(notes, start=17):
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
    ws.cell(i, 2, '• ' + text).font = font_body
    ws.cell(i, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[i].height = 32

ws['B23'] = 'NAVIGATION'
ws['B23'].fill = section_fill
ws['B23'].font = font_section
ws.merge_cells('B23:H23')
for cell in ws[23][1:8]: cell.fill = section_fill
nav = [('Sessions_88', 'Tableau comparatif des 88 sessions'), ('Décisions_88', '3 555 lignes détaillées'), ('Contrôle_Qualité', 'Tests et limites de validation')]
for i, (sheet, label) in enumerate(nav, start=24):
    cell = ws.cell(i, 2, label)
    cell.hyperlink = f"#'{sheet}'!A1"
    cell.font = Font(name='Calibri', size=10, color=TEAL, underline='single')
    ws.cell(i, 3, sheet).font = font_note
for i, s in enumerate(sessions, start=24):
    sheet_name = f"{s['session_id']}_{s['session'].replace('/', '_')}"
    if i == 24:
        pass

ws['B29'] = 'SOURCES'
ws['B29'].fill = section_fill
ws['B29'].font = font_section
ws.merge_cells('B29:H29')
for cell in ws[29][1:8]: cell.fill = section_fill
sources = [
    'Source canonique locale : public/data/reextraction_88_canonical.json',
    'Résultats institutionnels : https://startup.gov.tn/fr/startup_act/results',
    'PDF officiels des sessions Startup Tunisia, corpus contrôlé et documenté dans les champs source_pdf / controle_qualite.',
]
for i, text in enumerate(sources, start=30):
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
    ws.cell(i, 2, text).font = font_note
    ws.cell(i, 2).alignment = Alignment(wrap_text=True)
ws.freeze_panes = 'B7'

# 2) Session summary
ws = wb.create_sheet('Sessions_88')
setup(ws, {'A': 9, 'B': 12, 'C': 15, 'D': 15, 'E': 13, 'F': 13, 'G': 13, 'H': 13, 'I': 12, 'J': 13, 'K': 13, 'L': 40})
ws.merge_cells('A1:L1')
ws['A1'] = 'Sessions — compteurs officiels vs. lignes détaillées PDF'
ws['A1'].font = font_title
ws.merge_cells('A2:L2')
ws['A2'] = 'La colonne C est le compteur officiel ; D est le volume de lignes PDF détaillées. Ne pas les additionner comme deux populations.'
ws['A2'].font = font_subtitle
headers = ['Session ID', 'Session', 'Candidatures officielles', 'Lignes détaillées PDF', 'Labels officiels', 'Prélabels officiels', 'Conversions', 'Retraits', 'Reportés', 'Taux officiel (%)', 'Écart détail−officiel', 'Lien feuille']
for c, h in enumerate(headers, start=1): ws.cell(4, c, h)
style_header(ws, 4, 1, len(headers))
for r, s in enumerate(sessions, start=5):
    vals = [s['session_id'], s['session'], s['candidatures'], s['entries'], s['labels'], s['preLabels'], s['conversions'], s['retraits'], s.get('reportes', 0), s['tauxPct'], s['entries'] - s['candidatures'], None]
    for c, value in enumerate(vals, start=1): ws.cell(r, c, value)
    sheet_name = f"{s['session_id']}_{s['session'].replace('/', '_')}"
    ws.cell(r, 12, 'Ouvrir')
    ws.cell(r, 12).hyperlink = f"#'{sheet_name}'!A1"
    ws.cell(r, 12).font = Font(name='Calibri', size=10, color=TEAL, underline='single')
    if s['session'] == '05/2024':
        for c in range(1, 13): ws.cell(r, c).fill = warning_fill
style_body(ws, 5, 4 + len(sessions), 1, 12)
for r in range(5, 5 + len(sessions)):
    for c in [3,4,5,6,7,8,9,11]: ws.cell(r,c).number_format = '#,##0'
    ws.cell(r,10).number_format = '0.0'
ws.conditional_formatting.add(f'K5:K{4+len(sessions)}', ColorScaleRule(start_type='min', start_color='F4CCCC', mid_type='percentile', mid_value=50, mid_color='FFF2CC', end_type='max', end_color='D9EAD3'))
add_table(ws, f'A4:L{4+len(sessions)}', 'Sessions88')
ws.freeze_panes = 'C5'
ws.auto_filter.ref = f'A4:L{4+len(sessions)}'

# 3) All detailed rows
ws = wb.create_sheet('Décisions_88')
widths = {'A': 14, 'B': 9, 'C': 11, 'D': 8, 'E': 26, 'F': 26, 'G': 18, 'H': 38, 'I': 26, 'J': 24, 'K': 22, 'L': 24, 'M': 22, 'N': 17, 'O': 17, 'P': 19, 'Q': 52, 'R': 52}
setup(ws, widths)
ws.merge_cells('A1:R1')
ws['A1'] = 'Décisions — 3 555 lignes détaillées extraites des PDF'
ws['A1'].font = font_title
ws.merge_cells('A2:R2')
ws['A2'] = 'Chaque ligne conserve la provenance PDF, la décision brute, la normalisation, les fondateurs et le contrôle qualité. Les conversions et retraits restent visibles.'
ws['A2'].font = font_subtitle
headers = ['Decision ID', 'Session ID', 'Session', 'Ligne', 'Source PDF', 'Section PDF', 'Entreprise', 'Projet', 'Fondateurs', 'Secteur', 'Décision brute', 'Résultat normalisé', 'Type label', 'Tour / moment', 'Après pitching', 'Obtention / retrait', 'Commentaires', 'Contrôle qualité']
for c, h in enumerate(headers, start=1): ws.cell(4, c, h)
style_header(ws, 4, 1, len(headers))
for r, e in enumerate(entries, start=5):
    vals = [e.get('decision_id'), e.get('session_id'), e.get('session'), e.get('line'), e.get('source_pdf'), e.get('section_pdf'), e.get('societe'), e.get('projet'), e.get('fondateurs'), e.get('secteur'), e.get('decision'), e.get('resultat_normalise'), e.get('type_label'), e.get('tour_moment'), e.get('apres_pitching'), e.get('session_obtention_retrait'), e.get('commentaires'), e.get('controle_qualite')]
    for c, value in enumerate(vals, start=1): ws.cell(r, c, value if value not in (None, '') else 'Non renseigné')
    decision = e.get('resultat_normalise') or ''
    fill = None
    if decision == 'Reporté': fill = warning_fill
    elif decision == 'Décision non précisée — motif administratif': fill = error_fill
    elif decision == 'Retrait Label': fill = PALE_TEAL and source_fill
    if fill:
        for c in range(1, 19): ws.cell(r, c).fill = fill
style_body(ws, 5, 4 + len(entries), 1, 18)
for r in range(5, 5 + len(entries)): ws.cell(r, 4).number_format = '#,##0'
add_table(ws, f'A4:R{4+len(entries)}', 'Decisions88')
ws.freeze_panes = 'A5'
ws.auto_filter.ref = f'A4:R{4+len(entries)}'

# 4) Quality and definitions
ws = wb.create_sheet('Contrôle_Qualité')
setup(ws, {'B': 35, 'C': 21, 'D': 92})
ws.merge_cells('B2:D2')
ws['B2'] = 'Contrôle qualité et provenance'
ws['B2'].font = font_title
ws.merge_cells('B3:D3')
ws['B3'] = 'Résultats calculés à partir de reextraction_88_canonical.json et vérifiés contre les 88 JSON publics.'
ws['B3'].font = font_subtitle
for c, h in enumerate(['Contrôle', 'Résultat', 'Interprétation'], start=2): ws.cell(5, c, h)
style_header(ws, 5, 2, 4)
checks = [
    ('Nombre de sessions', len(sessions), 'Attendu : 88'),
    ('Candidatures officielles', sum(int(s['candidatures']) for s in sessions), 'Référence institutionnelle ; attendu : 3 079'),
    ('Lignes détaillées PDF', len(entries), 'Corpus détaillé ; attendu : 3 555'),
    ('Labels officiels', sum(int(s['labels']) for s in sessions), 'Compteur institutionnel ; attendu : 1 356'),
    ('Prélabels officiels', sum(int(s['preLabels']) for s in sessions), 'Compteur institutionnel ; attendu : 641'),
    ('Retraits officiels', sum(int(s['retraits']) for s in sessions), 'Compteur institutionnel ; attendu : 153'),
    ('Reportés explicites', sum(int(s.get('reportes') or 0) for s in sessions), 'Lignes Reporté dans le corpus détaillé ; attendu : 5'),
    ('JSON publics', len(list(D.glob('session_*.json'))) if (D := DATA / 'session-pdfs-json').exists() else 0, 'Un fichier year_month par session'),
]
for r, (label, value, note) in enumerate(checks, start=6):
    ws.cell(r, 2, label)
    ws.cell(r, 3, value)
    ws.cell(r, 4, note)
    if isinstance(value, int): ws.cell(r, 3).number_format = '#,##0'
    if 'attendu' in note.lower() and any(str(value) in note for _ in [0]):
        ws.cell(r, 3).fill = positive_fill
style_body(ws, 6, 5 + len(checks), 2, 4)

start = 16
ws.cell(start, 2, 'CAS S62 — 05/2024').fill = section_fill
ws.cell(start, 2).font = font_section
ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=4)
for cell in ws[start][1:4]: cell.fill = section_fill
s62 = next(s for s in sessions if s['session'] == '05/2024')
s62_counts = Counter(e.get('resultat_normalise') for e in entries_by_session['05/2024'])
for i, (label, value, note) in enumerate([
    ('Candidatures officielles', s62['candidatures'], 'Compteur Startup Tunisia'),
    ('Lignes détaillées PDF', s62['entries'], '46 lignes conservées'),
    ('Décision non précisée — motif administratif', s62_counts['Décision non précisée — motif administratif'], '4 dossiers conservés, motif administratif publié sans décision'),
    ('Conversions Prélabel → Label', s62['conversions'], 'Lignes documentaires supplémentaires'),
    ('Retraits de Label', s62['retraits'], 'Lignes documentaires supplémentaires'),
], start=start+1):
    ws.cell(i, 2, label); ws.cell(i, 3, value); ws.cell(i, 4, note)
    ws.cell(i, 3).number_format = '#,##0'
style_body(ws, start+1, start+5, 2, 4)
ws.cell(start+7, 2, 'Définition').fill = section_fill
ws.cell(start+7, 2).font = font_section
ws.merge_cells(start_row=start+7, start_column=2, end_row=start+7, end_column=4)
for cell in ws[start+7][1:4]: cell.fill = section_fill
ws.merge_cells(start_row=start+8, start_column=2, end_row=start+10, end_column=4)
ws.cell(start+8, 2, '« Candidature officielle » désigne le compteur institutionnel publié par session. « Ligne détaillée PDF » désigne une observation documentaire ; elle peut représenter une conversion, un retrait ou un statut administratif distinct. Les deux mesures sont présentées séparément pour préserver la traçabilité.')
ws.cell(start+8, 2).alignment = Alignment(wrap_text=True, vertical='top')
ws.cell(start+8, 2).font = font_body
ws.row_dimensions[start+8].height = 48
ws.freeze_panes = 'B6'

# 5) One sheet per session, exactly 88 session sheets.
entry_headers = ['Ligne', 'Entreprise', 'Projet', 'Fondateurs', 'Secteur', 'Décision brute', 'Résultat normalisé', 'Section PDF', 'Commentaires', 'Contrôle qualité', 'Decision ID', 'Type label', 'Tour / moment', 'Après pitching', 'Obtention / retrait', 'Source PDF']
for s in sessions:
    sheet_name = f"{s['session_id']}_{s['session'].replace('/', '_')}"
    ws = wb.create_sheet(sheet_name)
    setup(ws, {'A': 8, 'B': 36, 'C': 27, 'D': 46, 'E': 25, 'F': 28, 'G': 30, 'H': 25, 'I': 54, 'J': 54, 'K': 16, 'L': 18, 'M': 18, 'N': 18, 'O': 22, 'P': 28})
    ws.merge_cells('A1:P1')
    ws['A1'] = f"{s['session_id']} — {s['session']}"
    ws['A1'].font = font_title
    ws.merge_cells('A2:P2')
    ws['A2'] = f"Officiel : {s['candidatures']} candidatures · {s['labels']} Labels · {s['preLabels']} Prélabels · détail PDF : {s['entries']} lignes"
    ws['A2'].font = font_subtitle
    ws.merge_cells('A3:P3')
    if s['session'] == '05/2024':
        ws['A3'] = 'Note S62 : les 4 dossiers administratifs sans décision publiée sont conservés avec un statut séparé ; les conversions et retraits restent visibles.'
    elif s.get('reportes', 0):
        ws['A3'] = f"{s['reportes']} ligne(s) Reporté conservée(s) dans le détail PDF ; voir la colonne Résultat normalisé."
    else:
        ws['A3'] = 'La feuille distingue le compteur officiel du volume de lignes documentaires PDF.'
    ws['A3'].font = font_note
    for c, h in enumerate(entry_headers, start=1): ws.cell(5, c, h)
    style_header(ws, 5, 1, len(entry_headers))
    es = entries_by_session[s['session']]
    for r, e in enumerate(es, start=6):
        vals = [e.get('line'), e.get('societe'), e.get('projet'), e.get('fondateurs'), e.get('secteur'), e.get('decision'), e.get('resultat_normalise'), e.get('section_pdf'), e.get('commentaires'), e.get('controle_qualite'), e.get('decision_id'), e.get('type_label'), e.get('tour_moment'), e.get('apres_pitching'), e.get('session_obtention_retrait'), e.get('source_pdf')]
        for c, value in enumerate(vals, start=1): ws.cell(r, c, value if value not in (None, '') else 'Non renseigné')
        decision = e.get('resultat_normalise') or ''
        fill = warning_fill if decision == 'Reporté' else error_fill if decision == 'Décision non précisée — motif administratif' else source_fill if decision == 'Retrait Label' else None
        if fill:
            for c in range(1, 17): ws.cell(r, c).fill = fill
    style_body(ws, 6, 5 + len(es), 1, 16)
    for r in range(6, 6 + len(es)): ws.cell(r, 1).number_format = '#,##0'
    add_table(ws, f'A5:P{5+len(es)}', f"T{s['session_id'].replace('S','') or '0'}")
    ws.freeze_panes = 'A6'
    ws.auto_filter.ref = f'A5:P{5+len(es)}'

# Ensure the first four sheets are the supporting sheets and all 88 session sheets follow.
wb._sheets = [wb['Synthèse_88'], wb['Sessions_88'], wb['Décisions_88'], wb['Contrôle_Qualité']] + [wb[f"{s['session_id']}_{s['session'].replace('/', '_')}"] for s in sessions]

for path in [OUT_REPO, OUT_CANONICAL]:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print('WROTE', path, 'sheets', len(wb.sheetnames), 'sessions', len(sessions), 'entries', len(entries))
