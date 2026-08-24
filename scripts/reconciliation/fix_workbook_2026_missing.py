#!/usr/bin/env python3
from __future__ import annotations
import json
from pathlib import Path
from copy import copy
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / 'public' / 'data'
CANONICAL = json.loads((DATA/'reextraction_88_canonical.json').read_text(encoding='utf-8'))
BOOK = DATA/'Startup_Act_88_sessions_reextrait_corrige_2026-08-23.xlsx'
TARGETS = {'04/2026':'S85','05/2026':'S86','06/2026':'S87'}

def main():
    wb = load_workbook(BOOK)
    ws = wb['Décisions_88']
    existing = {str(ws.cell(r,1).value or '') for r in range(6, ws.max_row+1)}
    added = 0
    for e in CANONICAL['entries']:
        if e.get('session') not in TARGETS or e.get('decision_id') in existing: continue
        r = ws.max_row + 1
        vals = [e.get('decision_id'),e.get('session_id'),e.get('session'),e.get('line'),e.get('source_pdf'),e.get('section_pdf'),e.get('societe'),e.get('projet'),e.get('fondateurs'),e.get('secteur'),e.get('decision'),e.get('resultat_normalise'),e.get('type_label'),e.get('tour_moment'),e.get('apres_pitching'),e.get('session_obtention_retrait'),e.get('commentaires'),e.get('controle_qualite')]
        for c,v in enumerate(vals,1):
            cell = ws.cell(r,c,v if v not in (None,'') else 'Non renseigné')
            if r > 6:
                prev = ws.cell(r-1,c)
                if prev.has_style: cell._style = copy(prev._style)
        existing.add(e.get('decision_id')); added += 1
    # Update the session summary rows by matching the session label in column B or C.
    for row in range(1, ws.max_row+1):
        values = [str(ws.cell(row,c).value or '') for c in range(1, min(ws.max_column,10)+1)]
        for period, sid in TARGETS.items():
            if period in values or sid in values:
                for c in range(1, ws.max_column+1):
                    val = ws.cell(row,c).value
                    if val == 47 and period == '04/2026': ws.cell(row,c).value = 50
                    elif val == 42 and period == '05/2026': ws.cell(row,c).value = 48
                    elif val == 45 and period == '06/2026': ws.cell(row,c).value = 47
    wb.save(BOOK)
    print({'decisions_added': added, 'decisions_sheet_rows': ws.max_row})

if __name__ == '__main__': main()
