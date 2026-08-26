#!/usr/bin/env python3
"""Synchronise les exports après l’ajout des 11 lignes PDF 2026 manquantes."""
from __future__ import annotations

import csv
import json
import re
import shutil
from copy import copy
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "public" / "data"
CANONICAL = DATA / "reextraction_88_canonical.json"
SESSION_DIR = DATA / "session-pdfs-json"

TARGETS = {"01/2026": 42, "04/2026": 50, "05/2026": 48, "06/2026": 47}
SID = {"01/2026": "S82", "04/2026": "S85", "05/2026": "S86", "06/2026": "S87"}


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def dump_json(path: Path, value):
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def norm(s):
    return " ".join(str(s or "").casefold().split())


def entry_key(e):
    return norm(e.get("session")), norm(e.get("societe") or e.get("entreprise"))


def canonical_entries():
    c = load_json(CANONICAL)
    return c, c["entries"]


def sync_session_json(canonical):
    for session, target in TARGETS.items():
        path = SESSION_DIR / f"session_{session.split('/')[1]}_{session.split('/')[0]}.json"
        bundle = load_json(path)
        sid = SID[session]
        entries = [e for e in canonical["entries"] if e.get("session") == session]
        entries.sort(key=lambda e: (int(e.get("line") or 0), str(e.get("decision_id"))))
        assert len(entries) == target, (session, len(entries), target)
        bundle["entries"] = entries
        sd = bundle.setdefault("session_data", {})
        sd["entries_detaillees"] = target
        sd["candidatures_reexamen_pdf"] = target
        sd["candidatures_corrigees"] = target + int(sd.get("ajournes_hors_pdf") or 0)
        dump_json(path, bundle)


def csv_entry(e):
    return {
        "decision_id": e.get("decision_id"), "session_id": e.get("session_id"), "session": e.get("session"),
        "entreprise": e.get("societe"), "source_file": e.get("source_pdf"), "section": e.get("section_pdf"),
        "projet": e.get("projet"), "fondateurs": e.get("fondateurs"), "decision_brute": e.get("decision"),
        "resultat_normalise": e.get("resultat_normalise"), "type_decision": e.get("type_label"),
        "statut_demande": e.get("type_label"), "secteur": e.get("secteur"), "tour": e.get("tour_moment"),
        "apres_pitching": e.get("apres_pitching"), "obtention_retrait": e.get("session_obtention_retrait"),
        "commentaires": e.get("commentaires"), "controle_qualite": e.get("controle_qualite"),
    }


def sync_csvs(entries):
    comma_headers = ["decision_id","session_id","session","entreprise","source_file","section","projet","fondateurs","decision_brute","resultat_normalise","type_decision","statut_demande","secteur","tour","apres_pitching","obtention_retrait","commentaires","controle_qualite"]
    semicolon_headers = ["decision_id","session_id","session","line","source_pdf","section_pdf","societe","projet","fondateurs","secteur","decision","resultat_normalise","type_label","tour_moment","apres_pitching","session_obtention_retrait","commentaires","controle_qualite"]
    for name in ["database_88.csv", "database_entrees_brutes_88.csv", "database_entrees_brutes.csv"]:
        with (DATA / name).open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=comma_headers, lineterminator="\n")
            w.writeheader(); w.writerows(csv_entry(e) for e in entries)
    rows = []
    for e in entries:
        rows.append({"decision_id":e.get("decision_id"),"session_id":e.get("session_id"),"session":e.get("session"),"line":e.get("line"),"source_pdf":e.get("source_pdf"),"section_pdf":e.get("section_pdf"),"societe":e.get("societe"),"projet":e.get("projet"),"fondateurs":e.get("fondateurs"),"secteur":e.get("secteur"),"decision":e.get("decision"),"resultat_normalise":e.get("resultat_normalise"),"type_label":e.get("type_label"),"tour_moment":e.get("tour_moment"),"apres_pitching":e.get("apres_pitching"),"session_obtention_retrait":e.get("session_obtention_retrait"),"commentaires":e.get("commentaires"),"controle_qualite":e.get("controle_qualite")})
    for name in ["database_entrees_reextrait_88_corrige.csv"]:
        with (DATA / name).open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=semicolon_headers, delimiter=";", lineterminator="\n")
            w.writeheader(); w.writerows(rows)


def sync_sessions_tables(canonical):
    for name in ["sessions_table.json", "sessions_88.json", "startup_act_88_sessions.json"]:
        path = DATA / name
        if not path.exists(): continue
        obj = load_json(path)
        sessions = obj.get("sessions") if isinstance(obj, dict) else obj
        if not isinstance(sessions, list): continue
        for s in sessions:
            period = s.get("session")
            if period in TARGETS:
                s["entries"] = TARGETS[period]
                s["entries_detaillees"] = TARGETS[period]
                s["candidatures_reexamen_pdf"] = TARGETS[period]
                s["candidatures_corrigees"] = TARGETS[period] + int(s.get("ajournes_hors_pdf") or s.get("ajournes") or 0)
        if isinstance(obj, dict) and isinstance(obj.get("meta"), dict):
            meta = obj["meta"]
            if "totalCandidaturesReexamenPdf" in meta: meta["totalCandidaturesReexamenPdf"] = 3571
            if "totalCandidaturesCorrigees" in meta: meta["totalCandidaturesCorrigees"] = 3574
            if "ecartTotalPdfMoinsOfficiel" in meta: meta["ecartTotalPdfMoinsOfficiel"] = 3571 - 3079
            if "detailedEntries" in meta: meta["detailedEntries"] = 3571
            if "detailed_entries" in meta: meta["detailed_entries"] = 3571
        dump_json(path, obj)


def sync_dashboard_data():
    path = DATA / "dashboard_data.json"
    if not path.exists(): return
    obj = load_json(path)
    def walk(x):
        if isinstance(x, dict):
            if x.get("session") in TARGETS or x.get("period") in ["Avril 2026", "Mai 2026", "Juin 2026"]:
                p = x.get("session") or {"Avril 2026":"04/2026","Mai 2026":"05/2026","Juin 2026":"06/2026"}.get(x.get("period"))
                if p in TARGETS:
                    x["entries"] = TARGETS[p]; x["entries_detaillees"] = TARGETS[p]; x["candidatures_reexamen_pdf"] = TARGETS[p]
                    x["candidatures_corrigees"] = TARGETS[p] + int(x.get("ajournes_hors_pdf") or 0)
            for v in x.values(): walk(v)
        elif isinstance(x, list):
            for v in x: walk(v)
    walk(obj)
    dump_json(path, obj)


def sync_database_startups(entries):
    path = DATA / "database_startups_88.json"
    if not path.exists(): return
    arr = load_json(path)
    if not isinstance(arr, list): return
    existing = {(norm(x.get("session")), norm(x.get("societe") or x.get("nom"))) for x in arr}
    next_id = max((int(x.get("id") or 0) for x in arr), default=0) + 1
    for e in entries:
        k = (norm(e.get("session")), norm(e.get("societe")))
        if k in existing: continue
        if e.get("session") not in TARGETS: continue
        arr.append({"id": next_id, "decision_id":e.get("decision_id"), "session_id":e.get("session_id"), "session":e.get("session"), "nom":e.get("societe"), "societe":e.get("societe"), "projet":e.get("projet") or "Non renseigné", "secteur":e.get("secteur") or "Non renseigné", "anneeCreation":"2026", "labelDate":e.get("session") if e.get("resultat_normalise")=="Label accordé" else "", "siteWeb":"", "resume":"", "email":"", "telephone":"", "source":f"PDF détaillé — {e.get('source_pdf')}", "decision":e.get("resultat_normalise"), "flags":"", "founders":[n.strip() for n in str(e.get("fondateurs") or "").split(";") if n.strip()]})
        existing.add(k); next_id += 1
    dump_json(path, arr)


def sql_quote(v):
    if v is None or v == "": return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def sync_sql(entries, canonical):
    path = DATA / "startup_act_database.sql"
    text = path.read_text(encoding="utf-8")
    # Update metadata and session summary rows in the existing SQL.
    text = text.replace("INSERT INTO metadata(key,value) VALUES ('corrected_candidatures','3558');", "INSERT INTO metadata(key,value) VALUES ('corrected_candidatures','3574');")
    text = text.replace("INSERT INTO metadata(key,value) VALUES ('corrected_candidatures','3569');", "INSERT INTO metadata(key,value) VALUES ('corrected_candidatures','3574');")
    text = text.replace("INSERT INTO metadata(key,value) VALUES ('detailed_entries','3555');", "INSERT INTO metadata(key,value) VALUES ('detailed_entries','3571');")
    text = text.replace("INSERT INTO metadata(key,value) VALUES ('detailed_entries','3566');", "INSERT INTO metadata(key,value) VALUES ('detailed_entries','3571');")
    text = re.sub(r"(INSERT INTO sessions VALUES \('S85','04/2026','41','47','47')", r"\g<1>0", text) if False else text
    # Replace only the first summary columns for S85-S87 using exact known current rows.
    text = text.replace("INSERT INTO sessions VALUES ('S82','01/2026','31','37','37','0'", "INSERT INTO sessions VALUES ('S82','01/2026','31','42','42','0'")
    text = text.replace("INSERT INTO sessions VALUES ('S85','04/2026','41','47','47','0'", "INSERT INTO sessions VALUES ('S85','04/2026','41','50','50','0'")
    text = text.replace("INSERT INTO sessions VALUES ('S86','05/2026','40','42','42','0'", "INSERT INTO sessions VALUES ('S86','05/2026','40','48','48','0'")
    text = text.replace("INSERT INTO sessions VALUES ('S87','06/2026','40','45','45','0'", "INSERT INTO sessions VALUES ('S87','06/2026','40','47','47','0'")
    existing_ids = set(re.findall(r"INSERT INTO decisions VALUES \('([^']+)'", text))
    max_company = max([int(x) for x in re.findall(r"C(\d+)", text)] or [0])
    max_founder = max([int(x) for x in re.findall(r"F(\d+)", text)] or [0])
    additions = [e for e in entries if e.get("decision_id") not in existing_ids]
    company_ids = {}
    founder_ids = {}
    out = [text.rstrip(), ""]
    for i, e in enumerate(additions, start=1):
        cid = f"C{max_company+i:04d}"; company_ids[e["decision_id"]] = cid
        out.append("INSERT INTO companies VALUES (%s,%s,%s,%s);" % (sql_quote(cid),sql_quote(e.get("societe")),sql_quote(e.get("secteur")),sql_quote(e.get("session"))))
    founder_counter = 0
    for e in additions:
        for raw in [n.strip() for n in str(e.get("fondateurs") or "").split(";") if n.strip()]:
            founder_counter += 1; fid = f"F{max_founder+founder_counter:04d}"; founder_ids[(e["decision_id"],raw)] = fid
            out.append("INSERT INTO founders VALUES (%s,%s);" % (sql_quote(fid),sql_quote(raw)))
    for e in additions:
        cid = company_ids[e["decision_id"]]
        vals = [e.get("decision_id"),e.get("session_id"),cid,e.get("source_pdf"),e.get("section_pdf"),e.get("projet"),e.get("fondateurs"),e.get("decision"),e.get("resultat_normalise"),e.get("secteur"),e.get("tour_moment"),e.get("apres_pitching"),e.get("session_obtention_retrait"),e.get("commentaires"),e.get("controle_qualite")]
        out.append("INSERT INTO decisions VALUES (%s);" % ",".join(sql_quote(v) for v in vals))
    for e in additions:
        cid = company_ids[e["decision_id"]]
        for raw in [n.strip() for n in str(e.get("fondateurs") or "").split(";") if n.strip()]:
            out.append("INSERT INTO company_founders VALUES (%s);" % ",".join(sql_quote(v) for v in [e.get("decision_id"),cid,founder_ids[(e["decision_id"],raw)],e.get("session_id"),raw,"PDF 2026 audité"]))
    path.write_text("\n".join(out)+"\n", encoding="utf-8")


def sync_workbook(entries):
    path = DATA / "Startup_Act_88_sessions_reextrait_corrige_2026-08-23.xlsx"
    if not path.exists(): return
    wb = load_workbook(path)
    by_session = defaultdict(list)
    for e in entries: by_session[e.get("session")].append(e)
    for session, added in by_session.items():
        if session not in TARGETS: continue
        sheet = f"{SID[session]}_{session.replace('/', '_')}"
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        existing = {str(ws.cell(r, 2).value or "").strip().casefold() for r in range(6, ws.max_row+1)}
        for e in added:
            if str(e.get("societe") or "").strip().casefold() in existing: continue
            r = ws.max_row + 1
            vals = [e.get("line"),e.get("societe"),e.get("projet"),e.get("fondateurs"),e.get("secteur"),e.get("decision"),e.get("resultat_normalise"),e.get("section_pdf"),e.get("commentaires"),e.get("controle_qualite"),e.get("decision_id"),e.get("type_label"),e.get("tour_moment"),e.get("apres_pitching"),e.get("session_obtention_retrait"),e.get("source_pdf")]
            for c,v in enumerate(vals,1): ws.cell(r,c,v or "Non renseigné")
            existing.add(str(e.get("societe") or "").strip().casefold())
    # Update obvious global documentary totals without touching official counters.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace("3 555", "3 571").replace("3 558", "3 574").replace("3 566", "3 571").replace("3 569", "3 574")
                elif cell.value == 3555 or cell.value == 3566: cell.value = 3571
                elif cell.value == 3558 or cell.value == 3569: cell.value = 3574
    wb.save(path)


def main():
    canonical, entries = canonical_entries()
    counts = {s: sum(1 for e in entries if e.get("session") == s) for s in TARGETS}
    assert counts == TARGETS, counts
    sync_session_json(canonical)
    sync_csvs(entries)
    sync_sessions_tables(canonical)
    sync_dashboard_data()
    sync_database_startups(entries)
    sync_sql(entries, canonical)
    sync_workbook(entries)
    print(json.dumps({"canonical_entries": len(entries), "session_counts": counts, "total_with_ajournes": len(entries)+3}, ensure_ascii=False))

if __name__ == "__main__":
    main()
