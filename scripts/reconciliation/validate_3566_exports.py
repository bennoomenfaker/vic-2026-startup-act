#!/usr/bin/env python3
from __future__ import annotations
import csv
import json
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
D = ROOT/'public'/'data'
TARGETS = {'04/2026':50,'05/2026':48,'06/2026':47}
NAMES = ['Mathix Academy','SURUS','Tunisia transfert','Deep SaaS','Carbon Zero Tech','NFASS - نفس','FIXITECHPRO','Cuber','shopyia','Nvitee','Creedex']

def main():
    c=json.loads((D/'reextraction_88_canonical.json').read_text(encoding='utf-8'))
    assert len(c['entries'])==3566
    for s,n in TARGETS.items():
        assert sum(1 for e in c['entries'] if e.get('session')==s)==n
        p=D/f"session-pdfs-json/session_{s.split('/')[1]}_{s.split('/')[0]}.json"
        b=json.loads(p.read_text(encoding='utf-8'))
        assert len(b['entries'])==n
        assert len({e.get('societe') for e in b['entries']})==n
    for name in NAMES:
        assert any(e.get('societe')==name and e.get('session') in TARGETS for e in c['entries']), name
    for name in ['reextraction_88_canonical.sql','startup_act_database.sql','startup_act_database_88.sql','startup_act_database_reextrait_corrige_2026-08-23.sql']:
        count=sum(1 for line in (D/name).read_text(encoding='utf-8').splitlines() if line.startswith('INSERT INTO decisions'))
        assert count==3566, (name,count)
    for name in ['database_entrees_brutes_88.csv','database_entrees_reextrait_88_corrige.csv']:
        with (D/name).open(encoding='utf-8-sig',newline='') as f:
            assert sum(1 for _ in csv.DictReader(f))==3566, name
    wb=load_workbook(D/'Startup_Act_88_sessions_reextrait_corrige_2026-08-23.xlsx',read_only=True,data_only=True)
    assert len(wb.sheetnames)==92
    assert wb['Décisions_88'].max_row==3570
    print({'ok':True,'entries':3566,'corrected_with_ajournes':3569,'session_counts':TARGETS,'sql_decisions':3566,'xlsx_decisions_rows':3570})

if __name__=='__main__': main()
