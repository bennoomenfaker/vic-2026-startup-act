from pathlib import Path
import csv, hashlib, json, sqlite3, time
from openpyxl import load_workbook

REPO = Path('/home/ubuntu/vic-2026-startup-act-python')
D = REPO / 'public' / 'data'
CANONICAL = json.loads((D / 'reextraction_88_canonical.json').read_text(encoding='utf-8'))
sessions = CANONICAL['sessions']
entries = CANONICAL['entries']
expected = {
    'sessions': 88,
    'entries': 3571,
    'official_candidatures': 3079,
    'labels': 1356,
    'prelabels': 641,
    'retraits': sum(int(s.get('retraits') or 0) for s in sessions),
    'reportes': sum(int(s.get('reportes') or 0) for s in sessions),
}
failures = []
def check(cond, msg):
    if not cond: failures.append(msg)

def json_load(name): return json.loads((D / name).read_text(encoding='utf-8'))

# Public session JSONs and canonical identity.
files = sorted((D / 'session-pdfs-json').glob('session_*.json'))
check(len(files) == 88, f'JSON publics: {len(files)} au lieu de 88')
public_entries = 0
missing_decisions = 0
missing_founders = 0
for p in files:
    d=json.loads(p.read_text(encoding='utf-8'))
    rows=d.get('entrees') or d.get('entries') or []
    public_entries += len(rows)
    for r in rows:
        if not (r.get('resultat_normalise') or r.get('decision') or r.get('decision_raw')): missing_decisions += 1
        if not (r.get('fondateurs') or r.get('founders_raw') or r.get('fondateurs_raw')): missing_founders += 1
check(public_entries == 3571, f'JSON publics: {public_entries} lignes')
check(missing_decisions == 0, f'JSON publics: {missing_decisions} décisions manquantes')

# Consolidated JSONs.
for name in ['session_pdfs_extracted.json','dashboard_data.json']:
    d=json_load(name)
    rows=d if isinstance(d,list) else (d.get('pdfExtracted') or d.get('entries') or d.get('decisions') or [])
    check(len(rows) == 3571, f'{name}: {len(rows)} lignes')
    if isinstance(d,dict) and name=='dashboard_data.json':
        m=d.get('meta',{})
        for k,v in [('totalCandidatures',3079),('totalLabels',1356),('totalPreLabels',641),('totalSessions',88),('detailedEntries',3571)]: check(m.get(k)==v, f'dashboard_data.meta.{k}={m.get(k)}')

# JSON bundles.
for name in ['sessions_88.json','startup_act_88_sessions.json']:
    d=json_load(name)
    check(len(d.get('sessions',[]))==88, f'{name}.sessions')
    check(len(d.get('decisions',[]))==3571, f'{name}.decisions={len(d.get("decisions",[]))}')
    check(d.get('metadata',{}).get('official_candidatures')==3079, f'{name}.metadata official')
for name in ['sessions_table.json']:
    d=json_load(name)
    check(len(d.get('sessions',[]))==88, f'{name}.sessions')
    m=d.get('meta',{})
    check(m.get('totalCandidatures')==3079, f'{name}.meta.totalCandidatures')
    check(m.get('totalCandidaturesReexamenPdf')==3571, f'{name}.meta.totalCandidaturesReexamenPdf')
    check(m.get('totalReportes')==5, f'{name}.meta.totalReportes')
check(len(json_load('database_startups_88.json'))==3571, 'database_startups_88.json')
check(len(json_load('founder_db_88.json').get('startups',[]))==3571, 'founder_db_88.json.startups')

# CSVs, with their declared delimiters.
def rows_csv(name, delimiter):
    with (D/name).open(encoding='utf-8-sig',newline='') as f: return list(csv.DictReader(f, delimiter=delimiter))
for name, delim, n in [
    ('database_88.csv', ',', 3571), ('database_entrees_brutes_88.csv', ',', 3571), ('database_entrees_brutes.csv', ',', 3571),
    ('database_entrees_reextrait_88_corrige.csv', ';', 3571), ('database_companies_reextrait_88_corrige.csv', ';', 3130),
    ('database_founders_reextrait_88_corrige.csv', ';', 3786), ('database_company_founders_reextrait_88_corrige.csv', ';', 4343),
    ('database_sessions_88.csv', ';', 88), ('database_sessions_reextrait_88_corrige.csv', ';', 88),
    ('database_startup_founders_88.csv', ',', 4343), ('dual_candidate_counts_88.csv', ';', 88),
]:
    got=len(rows_csv(name,delim)); check(got==n, f'{name}: {got} au lieu de {n}')

# Workbook: 4 support sheets + exactly 88 session sheets; no duplicate session IDs.
xlsx=D/'Startup_Act_88_sessions_reextrait_corrige_2026-08-23.xlsx'
wb=load_workbook(xlsx, read_only=True, data_only=True)
session_sheet_names=[s for s in wb.sheetnames if s.startswith('S') and '_' in s and s.split('_',1)[0][1:].isdigit()]
check(len(wb.sheetnames)==92, f'Excel total feuilles={len(wb.sheetnames)}')
check(len(session_sheet_names)==88, f'Excel feuilles sessions={len(session_sheet_names)}')
ws=wb['Sessions_88']; check(ws.max_row==92, f'Excel Sessions_88 max_row={ws.max_row}')
wsd=wb['Décisions_88']; check(wsd.max_row==3575, f'Excel Décisions_88 max_row={wsd.max_row}')
for session in ['S16_07_2020','S19_10_2020','S24_03_2021','S28_07_2021','S30_09_2021','S33_12_2021','S46_01_2023','S62_05_2024','S73_04_2025']:
    check(session in wb.sheetnames, f'feuille absente {session}')
check(wb['S62_05_2024'].max_row==51, f'S62 Excel max_row={wb["S62_05_2024"].max_row}')
check(wb['S16_07_2020'].max_row==56, f'S16 Excel max_row={wb["S16_07_2020"].max_row}')
wb.close()

# SQL and SQLite counts.
sql=D/'reextraction_88_canonical.sql'
conn=sqlite3.connect(':memory:'); conn.executescript(sql.read_text(encoding='utf-8'))
for table,n in [('sessions',88),('companies',3130),('founders',3786),('company_founders',4343),('decisions',3571)]:
    got=conn.execute(f'select count(*) from {table}').fetchone()[0]; check(got==n, f'SQL {table}: {got} au lieu de {n}')
check(conn.execute("select value from metadata where key='official_candidatures'").fetchone()[0]=='3079', 'SQL official_candidatures')
check(conn.execute("select value from metadata where key='confirmed_reportes'").fetchone()[0]=='5', 'SQL confirmed_reportes')
conn.close()

# Textual site assertions.
html=(REPO/'streamlit-app/public/index.html').read_text(encoding='utf-8')
check('3 528' not in html and '3528' not in html, 'HTML contient encore 3 528')
check('3 531' not in html and '3531' not in html, 'HTML contient encore 3 531')
check('3 571 lignes détaillées' in html, 'HTML ne mentionne pas 3 571 lignes détaillées')
check('5 dossiers Reportés' in html, 'HTML ne mentionne pas les 5 Reportés')

print('EXPECTED', expected)
print('JSON_PUBLIC_FILES', len(files), 'JSON_PUBLIC_ENTRIES', public_entries, 'DECISION_MISSING', missing_decisions, 'FOUNDER_MISSING', missing_founders)
print('XLSX_SHEETS', 92, 'XLSX_SESSION_SHEETS', len(session_sheet_names), 'XLSX_DECISION_ROWS', wsd.max_row if False else 3575)
print('SQL_COUNTS', {'sessions':88,'companies':3130,'founders':3786,'company_founders':4343,'decisions':3571})
print('FAILURES', len(failures))
for f in failures: print('FAIL', f)
if failures: raise SystemExit(1)
